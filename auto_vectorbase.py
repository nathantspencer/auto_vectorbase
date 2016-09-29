from bs4 import BeautifulSoup
from openpyxl import load_workbook
import code
import requests
import re
import sys
from selenium import webdriver
import time

# vectorbaser
def vectorbaser(file_path):

    driver = webdriver.Firefox()

    # load up the worksheet, get the number of rows
    wb = load_workbook(filename=file_path, use_iterators=False)
    ws = wb.worksheets[0]

    num_rows = ws.get_highest_row()

    # iterate through target_ids, grab them
    target_ids = []
    for row in range(3, num_rows, 2):
        target_ids.append(ws.cell(row=row, column=1).value)

    # search all pages of drugs starting with that letter
    current_row_number = 3;
    wb.save(file_path);

    for target_id in target_ids:

        print(target_id)
        wb = load_workbook(filename=file_path, use_iterators=False)
        ws = wb.worksheets[0]

        # navigate to search page
        search_url = 'https://www.vectorbase.org/search/site/' + \
            target_id
        response = requests.get(search_url)
        soup = BeautifulSoup(response.text, 'lxml')

        # navigate to gene page
        gene_link = '/Gene/Summary/'
        pea_soup = soup.findAll('a', href=re.compile(gene_link))
        pea_soup_string = str(pea_soup.pop())
        link_ending = re.search('a href="(.*?)"', pea_soup_string)
        search_url = 'https://www.vectorbase.org' + link_ending.group(1)
        response = requests.get(search_url)
        soup = BeautifulSoup(response.text, 'lxml')

        # navigate to cellular component, grab info if it exists
        pea_soup = soup.findAll('a', {"title":"GO: Cellular component"})
        if len(pea_soup) > 0:
            pea_soup_string = str(pea_soup.pop())
            link_ending = re.search('href="(.*?)"', pea_soup_string)
            search_url = 'https://www.vectorbase.org' + link_ending.group(1)
            driver.get(search_url)
            time.sleep(3)
            source = driver.page_source
            cell_comp = re.findall('<td style="width:25%;text-align:left">([A-Za-z -_]+?)</td>', source)
            if len(cell_comp) > 0:
                cell_comp = cell_comp[0]
                ws.cell(row=current_row_number, column=3).value = cell_comp

        # navigate to biological process, grab info if it exists
        pea_soup = soup.findAll('a', {"title":"GO: Biological process"})
        if len(pea_soup) > 0:
            pea_soup_string = str(pea_soup.pop())
            link_ending = re.search('href="(.*?)"', pea_soup_string)
            search_url = 'https://www.vectorbase.org' + link_ending.group(1)
            driver.get(search_url)
            time.sleep(3)
            source = driver.page_source
            bio_proc = re.findall('<td style="width:25%;text-align:left">([A-Za-z -_]+?)</td>', source)
            if len(bio_proc) > 0:
                bio_proc = bio_proc[0]
                ws.cell(row=current_row_number, column=4).value = bio_proc

        # navigate to molecular function, grab info if it exists
        pea_soup = soup.findAll('a', {"title":"GO: Molecular function"})
        if len(pea_soup) > 0:
            pea_soup_string = str(pea_soup.pop())
            link_ending = re.search('href="(.*?)"', pea_soup_string)
            search_url = 'https://www.vectorbase.org' + link_ending.group(1)
            driver.get(search_url)
            time.sleep(3)
            source = driver.page_source
            mol_func = re.findall('<td style="width:25%;text-align:left">([A-Za-z -_]+?)</td>', source)
            if len(mol_func) > 0:
                mol_func = mol_func[0]
                ws.cell(row=current_row_number, column=5).value = mol_func

        # navigate to orthologues, grab info if it exists
        pea_soup = soup.findAll('a', {"title":"Orthologues"});
        if len(pea_soup) > 0:
            pea_soup_string = str(pea_soup.pop())
            link_ending = re.search('href="(.*?)"', pea_soup_string)
            search_url = 'https://www.vectorbase.org' + link_ending.group(1)
            driver.get(search_url)
            time.sleep(3)
            source = driver.page_source
            orthologues = re.findall('<td style="width:10%;text-align:left" class=" sorting_1">([a-zA-Z -_]+?)</td>', source)

            ortho_texts_wrapper = re.findall('<td style="width:15%;text-align:left">([a-zA-Z -_]+?)</td>', source)
            ortho_texts = []
            for wrapper in ortho_texts_wrapper:
                matches = re.findall('<span class="small">([a-zA-Z -_]+?)</span>', wrapper)
                ortho_texts.append(matches[0])

            ortho_dict = dict()
            for (orthologue, ortho_text) in zip(orthologues, ortho_texts):
                ortho_dict[orthologue] = ortho_text
            descriptionFound = False;
            description = 'No description available.';

            # check for glossina morsitans
            for orthologue in orthologues:
                if orthologue.split()[0] == "Glossina" and orthologue.split()[1] == "morsitans" and not ortho_dict[orthologue] == 'No description':
                    descriptionFound = True;
                    description = ortho_dict[orthologue];
                    break;

            # check for any glossina
            if not descriptionFound:
                for orthologue in orthologues:
                    if orthologue.split()[0] == "Glossina" and not ortho_dict[orthologue] == 'No description':
                        descriptionFound = True;
                        description = ortho_dict[orthologue];
                        break;

            # check for drosophila melanogaster
            if not descriptionFound:
                for orthologue in orthologues:
                    if orthologue == "Drosophila melanogaster" and not ortho_dict[orthologue] == 'No description':
                        descriptionFound = True;
                        description = ortho_dict[orthologue];
                        break;

            # check for musca domestica
            if not descriptionFound:
                for orthologue in orthologues:
                    if orthologue == "Musca domestica" and not ortho_dict[orthologue] == 'No description':
                        descriptionFound = True;
                        description = ortho_dict[orthologue];
                        break;

            # if we couldn't find a description, just admit defeat and write
            ws.cell(row=current_row_number, column=6).value = description

        # save, iterate to next row number
        wb.save(file_path)
        current_row_number += 2

    driver.Quit()


# help text and launch of vectorbaser
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('AUTO_VECTORBASE -- Written by Nathan Spencer 2016')
        print('Usage: python auto_vectorbase.py "path/to/excel/file.xlsx"')
    else:
        vectorbaser(sys.argv[1])
